import openpyxl as xl
from Tools.scripts.make_ctype import values
from openpyxl.chart import BarChart, Reference

# alias
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected = cell.value * 0.9
    corrected_price = sheet.cell(row, 4)
    corrected_price.value = corrected

Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e1')
wb.save('trans2.xlsx')












